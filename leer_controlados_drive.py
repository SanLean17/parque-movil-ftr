# ============================================================
# LEER CONTROLADOS DESDE GOOGLE DRIVE
# PARQUE MÓVIL FTR
#
# OBJETIVO:
# - Leer diariamente las planillas de Google Drive.
# - Detectar líneas individuales y múltiples.
# - Extraer FECHA / DOMINIO / INTERNO.
# - Mantener históricos.
# - Agregar nuevos controles sin perder los anteriores.
# - Evitar duplicados.
# ============================================================

from pathlib import Path
import io
import json
import os
import re

import pandas as pd

from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

from openpyxl import load_workbook

from configuracion_lineas import LINEAS_EMPRESAS


# ============================================================
# CONFIGURACIÓN
# ============================================================

CARPETA_CONTROLADOS = Path("controlados")

ID_CARPETA_DRIVE = (
    "1lS7Ybqr4Knej93BCdun4pDLVZX3ldv59"
)

GOOGLE_SERVICE_ACCOUNT_JSON = (
    "GOOGLE_SERVICE_ACCOUNT_JSON"
)

SCOPES = [
    "https://www.googleapis.com/auth/drive.readonly"
]


LINEAS_CONFIGURADAS = {
    str(linea)
    for linea in LINEAS_EMPRESAS.keys()
}


# ============================================================
# NORMALIZAR TEXTO
# ============================================================

def limpiar_texto(valor):

    if valor is None:
        return ""

    try:
        if pd.isna(valor):
            return ""
    except Exception:
        pass

    texto = str(valor)

    texto = re.sub(
        r"\s+",
        " ",
        texto
    )

    return texto.strip()


# ============================================================
# NORMALIZAR DOMINIO
# ============================================================

def normalizar_dominio(valor):

    texto = limpiar_texto(valor)

    if not texto:
        return ""

    return texto.upper()


# ============================================================
# NORMALIZAR INTERNO
# ============================================================

def normalizar_interno(valor):

    texto = limpiar_texto(valor)

    if not texto:
        return ""

    return texto


# ============================================================
# NORMALIZAR FECHA
# ============================================================

def normalizar_fecha(valor):

    if valor is None:
        return ""

    # --------------------------------------------------------
    # Fecha de Excel / Google Sheets
    # --------------------------------------------------------

    if hasattr(valor, "strftime"):

        try:
            return valor.strftime("%d/%m/%Y")
        except Exception:
            pass

    texto = limpiar_texto(valor)

    if not texto:
        return ""

    # --------------------------------------------------------
    # Si viene como timestamp:
    # 2026-08-21 00:00:00
    # --------------------------------------------------------

    match = re.match(
        r"^(\d{4})-(\d{1,2})-(\d{1,2})",
        texto
    )

    if match:

        anio = match.group(1)
        mes = match.group(2).zfill(2)
        dia = match.group(3).zfill(2)

        return f"{dia}/{mes}/{anio}"

    # --------------------------------------------------------
    # Si ya está DD/MM/YYYY
    # --------------------------------------------------------

    match = re.match(
        r"^(\d{1,2})/(\d{1,2})/(\d{4})$",
        texto
    )

    if match:

        dia = match.group(1).zfill(2)
        mes = match.group(2).zfill(2)
        anio = match.group(3)

        return f"{dia}/{mes}/{anio}"

    return texto


# ============================================================
# EXTRAER NÚMEROS DE LÍNEA
# ============================================================

def extraer_numeros_linea(texto):

    texto = limpiar_texto(texto)

    if not texto:
        return []

    encontrados = []

    # --------------------------------------------------------
    # Buscar números completos
    # --------------------------------------------------------

    numeros = re.findall(
        r"(?<!\d)(\d+)(?!\d)",
        texto
    )

    for numero in numeros:

        if numero in LINEAS_CONFIGURADAS:

            if numero not in encontrados:
                encontrados.append(numero)

    return encontrados


# ============================================================
# DETECTAR ENCABEZADO DE LÍNEA
# ============================================================

def detectar_lineas_en_celda(valor):

    texto = limpiar_texto(valor)

    if not texto:
        return []

    mayuscula = texto.upper()

    # --------------------------------------------------------
    # Encabezados explícitos:
    #
    # LINEA 95
    # LÍNEA 95
    # LINEA 63/113
    # LINEA 51/74/79/164/177
    # --------------------------------------------------------

    if (
        "LINEA" in mayuscula
        or "LÍNEA" in mayuscula
    ):

        return extraer_numeros_linea(texto)

    # --------------------------------------------------------
    # También permitimos:
    #
    # 95
    # 63
    # 113
    # --------------------------------------------------------

    if re.fullmatch(r"\d+", texto):

        if texto in LINEAS_CONFIGURADAS:
            return [texto]

    return []


# ============================================================
# IDENTIFICAR COLUMNAS
# ============================================================

def identificar_encabezados(
    ws,
    fila,
    columna_inicio,
    columna_fin
):

    resultado = {}

    for columna in range(
        columna_inicio,
        columna_fin + 1
    ):

        valor = limpiar_texto(
            ws.cell(
                fila,
                columna
            ).value
        ).lower()

        if not valor:
            continue

        normalizado = (
            valor
            .replace("_", "")
            .replace("-", "")
            .replace(" ", "")
        )

        # ----------------------------------------------------
        # FECHA
        # ----------------------------------------------------

        if (
            "fecha" in normalizado
            and "fecha" not in resultado
        ):

            resultado["fecha"] = columna

        # ----------------------------------------------------
        # DOMINIO
        # ----------------------------------------------------

        elif (
            "dominio" in normalizado
            and "dominio" not in resultado
        ):

            resultado["dominio"] = columna

        # ----------------------------------------------------
        # INTERNO
        # ----------------------------------------------------

        elif (
            "interno" in normalizado
            and "interno" not in resultado
        ):

            resultado["interno"] = columna

    return resultado


# ============================================================
# DETERMINAR LÍMITES DE UN BLOQUE HORIZONTAL
# ============================================================

def determinar_limites_bloque(
    ws,
    fila,
    columna_linea
):

    # --------------------------------------------------------
    # Buscar hasta 20 columnas hacia la derecha.
    #
    # Normalmente el bloque es:
    #
    # FECHA | DOMINIO | INTERNO
    #
    # y después hay una columna separadora.
    # --------------------------------------------------------

    inicio = columna_linea

    fin = min(
        ws.max_column,
        columna_linea + 20
    )

    return inicio, fin


# ============================================================
# BUSCAR BLOQUES
# ============================================================

def encontrar_bloques(ws, nombre_archivo=""):

    candidatos = []

    # ========================================================
    # PASO 1
    # Buscar todos los encabezados de línea.
    # ========================================================

    for fila in range(
        1,
        ws.max_row + 1
    ):

        for columna in range(
            1,
            ws.max_column + 1
        ):

            valor = ws.cell(
                fila,
                columna
            ).value

            lineas = detectar_lineas_en_celda(
                valor
            )

            if not lineas:
                continue

            candidatos.append(
                {
                    "fila": fila,
                    "columna": columna,
                    "lineas": lineas
                }
            )

    # ========================================================
    # PASO 2
    # Buscar los encabezados FECHA / DOMINIO / INTERNO.
    # ========================================================

    bloques = []

    for candidato in candidatos:

        fila_linea = candidato["fila"]
        columna_linea = candidato["columna"]
        lineas = candidato["lineas"]

        inicio_columna, fin_columna = (
            determinar_limites_bloque(
                ws,
                fila_linea,
                columna_linea
            )
        )

        encabezados = None
        fila_encabezados = None

        # ----------------------------------------------------
        # Buscar durante las siguientes 10 filas
        # ----------------------------------------------------

        for fila in range(
            fila_linea + 1,
            min(
                ws.max_row + 1,
                fila_linea + 11
            )
        ):

            posibles = identificar_encabezados(
                ws,
                fila,
                inicio_columna,
                fin_columna
            )

            if (
                "fecha" in posibles
                and "dominio" in posibles
                and "interno" in posibles
            ):

                encabezados = posibles
                fila_encabezados = fila

                break

        if encabezados is None:
            continue

        # ====================================================
        # Si hay varias líneas en un mismo encabezado:
        #
        # LINEA 63/113
        #
        # primero buscamos otros bloques independientes.
        # ====================================================

        if len(lineas) == 1:

            bloques.append(
                {
                    "linea": lineas[0],
                    "fila_linea": fila_linea,
                    "columna_linea": columna_linea,
                    "fila_encabezados": fila_encabezados,
                    "col_fecha": encabezados["fecha"],
                    "col_dominio": encabezados["dominio"],
                    "col_interno": encabezados["interno"]
                }
            )

        else:

            # ------------------------------------------------
            # Caso especial:
            #
            # LINEA 63/113
            #
            # Si existen varias líneas pero un único bloque
            # horizontal, intentamos encontrar bloques
            # adicionales desplazándonos hacia la derecha.
            # ------------------------------------------------

            bloques_encontrados = []

            columnas_usadas = set()

            for posible_columna in range(
                columna_linea,
                ws.max_column + 1
            ):

                posibles = identificar_encabezados(
                    ws,
                    fila_encabezados,
                    posible_columna,
                    min(
                        ws.max_column,
                        posible_columna + 4
                    )
                )

                if not (
                    "fecha" in posibles
                    and "dominio" in posibles
                    and "interno" in posibles
                ):
                    continue

                clave = (
                    posibles["fecha"],
                    posibles["dominio"],
                    posibles["interno"]
                )

                if clave in columnas_usadas:
                    continue

                columnas_usadas.add(clave)

                bloques_encontrados.append(
                    {
                        "fila_linea": fila_linea,
                        "columna_linea": posible_columna,
                        "fila_encabezados": fila_encabezados,
                        "col_fecha": posibles["fecha"],
                        "col_dominio": posibles["dominio"],
                        "col_interno": posibles["interno"]
                    }
                )

            # ------------------------------------------------
            # Si encontramos varios bloques, asignarlos en
            # el mismo orden de las líneas del encabezado.
            # ------------------------------------------------

            if len(bloques_encontrados) >= len(lineas):

                bloques_encontrados.sort(
                    key=lambda x: (
                        x["fila_linea"],
                        x["columna_linea"]
                    )
                )

                for indice, linea in enumerate(lineas):

                    bloque = bloques_encontrados[
                        indice
                    ]

                    bloques.append(
                        {
                            "linea": linea,
                            **bloque
                        }
                    )

            else:

                # ------------------------------------------------
                # Si solamente existe un bloque y el encabezado
                # contiene varias líneas, lo asociamos a la
                # primera línea. Los demás bloques deberán tener
                # su propio encabezado.
                # ------------------------------------------------

                bloques.append(
                    {
                        "linea": lineas[0],
                        "fila_linea": fila_linea,
                        "columna_linea": columna_linea,
                        "fila_encabezados": fila_encabezados,
                        "col_fecha": encabezados["fecha"],
                        "col_dominio": encabezados["dominio"],
                        "col_interno": encabezados["interno"]
                    }
                )

    # ========================================================
    # ELIMINAR DUPLICADOS
    # ========================================================

    unicos = []
    claves = set()

    for bloque in bloques:

        clave = (
            bloque["linea"],
            bloque["fila_linea"],
            bloque["columna_linea"],
            bloque["fila_encabezados"],
            bloque["col_fecha"],
            bloque["col_dominio"],
            bloque["col_interno"]
        )

        if clave in claves:
            continue

        claves.add(clave)
        unicos.append(bloque)

    # ========================================================
    # ORDENAR
    # ========================================================

    unicos.sort(
        key=lambda x: (
            int(x["linea"]),
            x["fila_linea"],
            x["columna_linea"]
        )
    )

    return unicos


# ============================================================
# EXTRAER DATOS DE UN BLOQUE
# ============================================================

def extraer_datos_bloque(
    ws,
    bloque
):

    linea = str(
        bloque["linea"]
    )

    fila_inicio = (
        bloque["fila_encabezados"] + 1
    )

    col_fecha = bloque["col_fecha"]
    col_dominio = bloque["col_dominio"]
    col_interno = bloque["col_interno"]

    registros = []

    filas_vacias = 0

    # ========================================================
    # LEER FILAS
    # ========================================================

    for fila in range(
        fila_inicio,
        ws.max_row + 1
    ):

        # ----------------------------------------------------
        # Si encontramos un nuevo encabezado de línea
        # en otra fila, termina el bloque.
        # ----------------------------------------------------

        encontro_otro_encabezado = False

        if fila > fila_inicio:

            for columna in range(
                1,
                ws.max_column + 1
            ):

                valor = ws.cell(
                    fila,
                    columna
                ).value

                lineas = detectar_lineas_en_celda(
                    valor
                )

                if lineas:

                    encontro_otro_encabezado = True
                    break

        if encontro_otro_encabezado:
            break

        # ----------------------------------------------------
        # Leer datos
        # ----------------------------------------------------

        fecha = ws.cell(
            fila,
            col_fecha
        ).value

        dominio = ws.cell(
            fila,
            col_dominio
        ).value

        interno = ws.cell(
            fila,
            col_interno
        ).value

        fecha = normalizar_fecha(
            fecha
        )

        dominio = normalizar_dominio(
            dominio
        )

        interno = normalizar_interno(
            interno
        )

        # ----------------------------------------------------
        # Verificar si hay vehículo
        # ----------------------------------------------------

        if dominio or interno:

            registros.append(
                {
                    "linea": linea,
                    "fecha": fecha,
                    "dominio": dominio,
                    "interno": interno
                }
            )

            filas_vacias = 0

        else:

            filas_vacias += 1

            if filas_vacias >= 3:
                break

    # ========================================================
    # DATAFRAME
    # ========================================================

    columnas = [
        "linea",
        "fecha",
        "dominio",
        "interno"
    ]

    if not registros:

        return pd.DataFrame(
            columns=columnas
        )

    df = pd.DataFrame(
        registros
    )

    # --------------------------------------------------------
    # Eliminar filas sin datos
    # --------------------------------------------------------

    df = df[
        (
            df["dominio"].astype(str).str.strip()
            != ""
        )
        |
        (
            df["interno"].astype(str).str.strip()
            != ""
        )
    ].copy()

    return df


# ============================================================
# CONECTAR CON GOOGLE DRIVE
# ============================================================

def conectar_drive():

    contenido = os.environ.get(
        GOOGLE_SERVICE_ACCOUNT_JSON
    )

    if not contenido:

        raise RuntimeError(
            "No se encontró el Secret "
            "GOOGLE_SERVICE_ACCOUNT_JSON."
        )

    try:

        datos = json.loads(
            contenido
        )

    except Exception as error:

        raise RuntimeError(
            "El Secret "
            "GOOGLE_SERVICE_ACCOUNT_JSON "
            "no contiene un JSON válido."
        ) from error

    credenciales = (
        service_account
        .Credentials
        .from_service_account_info(
            datos,
            scopes=SCOPES
        )
    )

    return build(
        "drive",
        "v3",
        credentials=credenciales
    )


# ============================================================
# LISTAR TODOS LOS ARCHIVOS DE DRIVE
# ============================================================

def listar_archivos_drive(
    drive
):

    consulta = (
        f"'{ID_CARPETA_DRIVE}' "
        "in parents "
        "and trashed = false"
    )

    archivos = []

    page_token = None

    while True:

        respuesta = (
            drive.files()
            .list(
                q=consulta,
                fields=(
                    "nextPageToken,"
                    "files("
                    "id,"
                    "name,"
                    "mimeType"
                    ")"
                ),
                pageSize=1000,
                pageToken=page_token
            )
            .execute()
        )

        archivos.extend(
            respuesta.get(
                "files",
                []
            )
        )

        page_token = (
            respuesta.get(
                "nextPageToken"
            )
        )

        if not page_token:
            break

    return archivos


# ============================================================
# DESCARGAR ARCHIVO
# ============================================================

def descargar_archivo(
    drive,
    archivo
):

    file_id = archivo["id"]

    mime_type = archivo[
        "mimeType"
    ]

    if mime_type == (
        "application/vnd.google-apps.spreadsheet"
    ):

        request = (
            drive.files()
            .export_media(
                fileId=file_id,
                mimeType=(
                    "application/vnd.openxmlformats-"
                    "officedocument.spreadsheetml.sheet"
                )
            )
        )

    else:

        request = (
            drive.files()
            .get_media(
                fileId=file_id
            )
        )

    buffer = io.BytesIO()

    downloader = (
        MediaIoBaseDownload(
            buffer,
            request
        )
    )

    terminado = False

    while not terminado:

        _, terminado = (
            downloader.next_chunk()
        )

    buffer.seek(0)

    return buffer


# ============================================================
# PROCESAR ARCHIVO
# ============================================================

def procesar_archivo(
    drive,
    archivo
):

    nombre = limpiar_texto(
        archivo.get(
            "name",
            ""
        )
    )

    print()
    print("=" * 70)
    print(
        f"ARCHIVO DRIVE: {nombre}"
    )
    print("=" * 70)

    # ========================================================
    # DESCARGAR
    # ========================================================

    try:

        contenido = descargar_archivo(
            drive,
            archivo
        )

    except Exception as error:

        print(
            f"ERROR descargando "
            f"{nombre}: {error}"
        )

        return {}

    # ========================================================
    # ABRIR EXCEL
    # ========================================================

    try:

        wb = load_workbook(
            contenido,
            data_only=True
        )

    except Exception as error:

        print(
            f"ERROR abriendo "
            f"{nombre}: {error}"
        )

        return {}

    resultados = {}

    # ========================================================
    # RECORRER HOJAS
    # ========================================================

    for ws in wb.worksheets:

        print()
        print(
            f"Hoja: {ws.title}"
        )

        bloques = encontrar_bloques(
            ws,
            nombre
        )

        print(
            f"Bloques encontrados: "
            f"{len(bloques)}"
        )

        if not bloques:
            continue

        for bloque in bloques:

            linea = str(
                bloque["linea"]
            )

            print()
            print(
                f"Procesando Línea {linea}..."
            )

            print(
                f"  Fila encabezado: "
                f"{bloque['fila_linea']}"
            )

            print(
                f"  Columna bloque: "
                f"{bloque['columna_linea']}"
            )

            print(
                f"  Fila columnas: "
                f"{bloque['fila_encabezados']}"
            )

            df = extraer_datos_bloque(
                ws,
                bloque
            )

            cantidad = len(df)

            print(
                f"  Línea {linea}: "
                f"{cantidad} registros encontrados."
            )

            if df.empty:
                continue

            if linea not in resultados:
                resultados[linea] = []

            resultados[linea].append(
                df
            )

    return resultados


# ============================================================
# LEER CSV ANTERIOR
# ============================================================

def leer_csv_anterior(
    archivo
):

    if not archivo.exists():
        return pd.DataFrame()

    try:

        df = pd.read_csv(
            archivo,
            dtype=str,
            encoding="utf-8-sig"
        )

    except Exception as error:

        print(
            f"ADVERTENCIA leyendo "
            f"{archivo}: {error}"
        )

        return pd.DataFrame()

    return df


# ============================================================
# PREPARAR DATAFRAME
# ============================================================

def preparar_dataframe(
    df
):

    columnas = [
        "linea",
        "fecha",
        "dominio",
        "interno"
    ]

    for columna in columnas:

        if columna not in df.columns:
            df[columna] = ""

    df = df[
        columnas
    ].copy()

    df["linea"] = (
        df["linea"]
        .astype(str)
        .str.strip()
    )

    df["fecha"] = (
        df["fecha"]
        .astype(str)
        .str.strip()
    )

    df["dominio"] = (
        df["dominio"]
        .astype(str)
        .str.strip()
        .str.upper()
    )

    df["interno"] = (
        df["interno"]
        .astype(str)
        .str.strip()
    )

    # --------------------------------------------------------
    # Quitar NaN convertidos a texto
    # --------------------------------------------------------

    for columna in columnas:

        df.loc[
            df[columna].isin(
                [
                    "nan",
                    "None",
                    "NaT"
                ]
            ),
            columna
        ] = ""

    # --------------------------------------------------------
    # Solo registros reales
    # --------------------------------------------------------

    df = df[
        (
            df["dominio"] != ""
        )
        |
        (
            df["interno"] != ""
        )
    ].copy()

    return df


# ============================================================
# UNIR Y GUARDAR
# ============================================================

def guardar_resultados(
    resultados
):

    CARPETA_CONTROLADOS.mkdir(
        parents=True,
        exist_ok=True
    )

    for linea, dataframes in resultados.items():

        linea = str(linea)

        if not dataframes:
            continue

        # ====================================================
        # DATOS NUEVOS
        # ====================================================

        df_nuevo = pd.concat(
            dataframes,
            ignore_index=True
        )

        df_nuevo = preparar_dataframe(
            df_nuevo
        )

        if df_nuevo.empty:
            continue

        archivo_salida = (
            CARPETA_CONTROLADOS
            / f"linea{linea}.csv"
        )

        # ====================================================
        # DATOS ANTERIORES
        # ====================================================

        df_anterior = leer_csv_anterior(
            archivo_salida
        )

        if not df_anterior.empty:

            df_anterior = preparar_dataframe(
                df_anterior
            )

            df_nuevo = pd.concat(
                [
                    df_anterior,
                    df_nuevo
                ],
                ignore_index=True
            )

        # ====================================================
        # NORMALIZAR
        # ====================================================

        df_nuevo = preparar_dataframe(
            df_nuevo
        )

        # ====================================================
        # DEDUPLICAR
        #
        # Si existe la misma patente varias veces,
        # conservamos el último registro.
        # ====================================================

        df_con_dominio = df_nuevo[
            df_nuevo["dominio"] != ""
        ].copy()

        df_sin_dominio = df_nuevo[
            df_nuevo["dominio"] == ""
        ].copy()

        if not df_con_dominio.empty:

            df_con_dominio = (
                df_con_dominio
                .drop_duplicates(
                    subset=[
                        "linea",
                        "dominio"
                    ],
                    keep="last"
                )
            )

        if not df_sin_dominio.empty:

            df_sin_dominio = (
                df_sin_dominio
                .drop_duplicates(
                    subset=[
                        "linea",
                        "interno"
                    ],
                    keep="last"
                )
            )

        df_final = pd.concat(
            [
                df_con_dominio,
                df_sin_dominio
            ],
            ignore_index=True
        )

        # ====================================================
        # ORDENAR
        # ====================================================

        df_final = df_final.sort_values(
            by=[
                "linea",
                "interno",
                "dominio"
            ],
            kind="stable"
        )

        # ====================================================
        # GUARDAR
        # ====================================================

        df_final.to_csv(
            archivo_salida,
            index=False,
            encoding="utf-8-sig"
        )

        print()
        print(
            f"GUARDADO CONTROLADOS "
            f"LÍNEA {linea}: "
            f"{len(df_final)} registros"
        )

        print(
            f"Archivo: {archivo_salida}"
        )


# ============================================================
# PROGRAMA PRINCIPAL
# ============================================================

def main():

    print()
    print("=" * 70)
    print(
        "LECTOR DE CONTROLADOS DESDE GOOGLE DRIVE"
    )
    print("=" * 70)

    print(
        f"Carpeta Drive: "
        f"{ID_CARPETA_DRIVE}"
    )

    print()

    # ========================================================
    # CONECTAR
    # ========================================================

    drive = conectar_drive()

    print(
        "Conexión con Google Drive: OK"
    )

    # ========================================================
    # LISTAR
    # ========================================================

    archivos = listar_archivos_drive(
        drive
    )

    print(
        f"Archivos encontrados: "
        f"{len(archivos)}"
    )

    # ========================================================
    # PROCESAR
    # ========================================================

    resultados_totales = {}

    for archivo in archivos:

        nombre = limpiar_texto(
            archivo.get(
                "name",
                ""
            )
        )

        nombre_mayuscula = (
            nombre.upper()
        )

        extensiones_validas = (
            ".XLSX",
            ".XLS",
            ".CSV",
            ".ODS"
        )

        es_planilla = (
            nombre_mayuscula.endswith(
                extensiones_validas
            )
            or
            archivo.get("mimeType")
            == (
                "application/vnd.google-apps.spreadsheet"
            )
        )

        if not es_planilla:
            continue

        resultados = procesar_archivo(
            drive,
            archivo
        )

        for linea, dataframes in resultados.items():

            if linea not in resultados_totales:

                resultados_totales[
                    linea
                ] = []

            resultados_totales[
                linea
            ].extend(
                dataframes
            )

    # ========================================================
    # GUARDAR
    # ========================================================

    print()
    print("=" * 70)
    print(
        "GUARDANDO CONTROLADOS"
    )
    print("=" * 70)

    guardar_resultados(
        resultados_totales
    )

    # ========================================================
    # RESUMEN
    # ========================================================

    print()
    print("=" * 70)
    print(
        "RESUMEN DE LECTURA DE DRIVE"
    )
    print("=" * 70)

    for linea in sorted(
        resultados_totales.keys(),
        key=lambda x: int(x)
    ):

        cantidad = sum(
            len(df)
            for df in resultados_totales[
                linea
            ]
        )

        print(
            f"Línea {linea}: "
            f"{cantidad} registros leídos desde Drive"
        )

    print("=" * 70)
    print(
        "LECTURA DE DRIVE COMPLETADA"
    )
    print("=" * 70)


# ============================================================
# EJECUCIÓN
# ============================================================

if __name__ == "__main__":
    main()
